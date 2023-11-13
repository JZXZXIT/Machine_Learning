# 若希望打包为exe文件，请在终端键入：pyinstaller -D main.py --noconsole

from MyClasses import KNN
import openpyxl
import pandas
import numpy
import matplotlib.pyplot as plt
import PySimpleGUI

def 更新数据():
    数据表 = openpyxl.load_workbook("./身高体重数据.xlsx")
    数据 = 数据表['Sheet1']
    行数 = 数据.max_row
    列数 = 数据.max_column
    标签 = []
    训练数据 = []
    数据集合 = []
    for 行 in range(2, 行数):
        标签.append(数据.cell(row=行, column=1).value)
        i = [数据.cell(row=行, column=2).value, 数据.cell(row=行, column=3).value]
        if 数据.cell(row=行, column=1).value == '男':
            a = 0
        else:
            a = 1
        ii = [数据.cell(row=行, column=2).value, 数据.cell(row=行, column=3).value, a]
        训练数据.append(i)
        数据集合.append(ii)
    return 训练数据, 标签, 数据表, 数据集合

# 后面没有调用该函数，使用时直接调用即可
# 该函数的目的仅为数据可视化，即前期验证模型是否可行
def 绘图():  # 男为0，女为1
    数据集合 = 更新数据()[3]
    df = pandas.DataFrame(数据集合, columns=['hight', 'wight', 'gender'])
    df.plot.scatter(x='hight', y='wight', c='gender', colormap='viridis')
    plt.show()

def 预测(值, 训练数据, 标签):
    值[0] = float(值[0])
    值[1] = float(值[1])
    输入数据 = [值[0], 值[1]]
    输入数据 = numpy.array(输入数据)
    训练数据 = numpy.array(训练数据)
    预测性别 = KNN(输入数据, 训练数据, 标签, 3)
    return 预测性别, 值

def 增添信息(反馈, 值, 预测性别, 数据表, 训练数据):
    数据 = 数据表['Sheet1']
    if 反馈 == "Yes":
        性别 = 预测性别
    elif 反馈 == "No":
        if 预测性别 == "男":
            性别 = "女"
        else:
            性别 = "男"
    else:
        return
    是否重复 = False
    for i in 训练数据:
        if i == [值[0], 值[1]]:
            是否重复 = True
            break
    if not 是否重复:
        输入列表 = [性别, 值[0], 值[1]]
        数据.append(输入列表)
        数据表.save("./身高体重数据.xlsx")


if __name__ == "__main__":
    ### 初始化布局
    布局 = [
        [PySimpleGUI.Text('请输入您的身高(cm)'), PySimpleGUI.InputText('')],
        [PySimpleGUI.Text('请输入您的体重(Kg)'), PySimpleGUI.InputText('')],
        [PySimpleGUI.Button('确定')]
    ]
    PySimpleGUI.theme('DarkTeal1')
    窗口 = PySimpleGUI.Window('性别预测', 布局, element_justification='c')

    ### 应用GUI
    while True:
        事件, 值 = 窗口.read()
        if 事件 == None:
             break
        if 事件 == '确定':
            训练数据, 标签, 数据表, a = 更新数据()
            预测性别, 输入值 = 预测(值, 训练数据, 标签)
            反馈 = PySimpleGUI.popup_yes_no(f"预测出您的性别为\n{预测性别}\n\n请问是否正确", title='预测性别')
            增添信息(反馈, 输入值, 预测性别, 数据表, 训练数据)

    窗口.close()